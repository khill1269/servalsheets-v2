#!/usr/bin/env python3
"""
Generate reports from SpreadsheetBench results.

Outputs:
  1. JSON summary (already created by run_benchmark.py)
  2. XLSX spreadsheet with per-instruction results and pivot tables
  3. Interactive HTML dashboard with charts

Usage:
    python generate_report.py --results-dir results/track_a_single_claude-sonnet-4_20260322_120000/
    python generate_report.py --compare results/track_a_*/  # Compare multiple runs
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


# ============================================================================
# Known competitor scores (from SpreadsheetBench paper + leaderboard)
# ============================================================================

COMPETITOR_SCORES = {
    # Format: model_name: { setting: { "soft": float, "hard": float } }
    # Source: SpreadsheetBench paper Table 2 (NeurIPS 2024)
    "GPT-4o (single)": {"soft": 0.441, "hard": 0.381},
    "GPT-4-Turbo (single)": {"soft": 0.381, "hard": 0.322},
    "Claude-3.5-Sonnet (single)": {"soft": 0.386, "hard": 0.329},
    "GPT-4o (multi-row-react-exec)": {"soft": 0.542, "hard": 0.480},
    "Claude-3.5-Sonnet (multi-row-react-exec)": {"soft": 0.494, "hard": 0.428},
    "Llama-3.1-70B (single)": {"soft": 0.213, "hard": 0.170},
    "Qwen2-72B (single)": {"soft": 0.241, "hard": 0.194},
    "DeepSeek-V2.5 (single)": {"soft": 0.285, "hard": 0.238},
}


# ============================================================================
# XLSX Report Generator
# ============================================================================

def generate_xlsx_report(results_data, output_path):
    """Generate a detailed XLSX report with per-instruction results and summary sheets."""

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary = results_data.get("summary", {})
    scores = summary.get("scores", {})
    overall = scores.get("overall", {})
    cell_level = scores.get("cell_level", {})
    sheet_level = scores.get("sheet_level", {})
    cost = summary.get("cost", {})

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True)
    metric_font = Font(name="Calibri", size=11)
    pct_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    # Title
    ws_summary["A1"] = "SpreadsheetBench Results"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")

    # Run metadata
    rows = [
        ("Run ID", summary.get("run_id", "")),
        ("Track", summary.get("track", "")),
        ("Setting", summary.get("setting", "")),
        ("Model", summary.get("model", "")),
        ("Timestamp", summary.get("timestamp", "")),
        ("Dataset Size", summary.get("dataset_size", 0)),
        ("Completed", summary.get("completed", 0)),
        ("Errors", summary.get("errors", 0)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = Font(bold=True)
        ws_summary[f"B{i}"] = value

    # Scores section
    row = 12
    ws_summary[f"A{row}"] = "Scores"
    ws_summary[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    score_headers = ["Category", "Count", "Soft Score", "Hard Score"]
    for col, header in enumerate(score_headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    score_rows = [
        ("Overall", summary.get("dataset_size", 0),
         overall.get("soft_avg", 0), overall.get("hard_avg", 0)),
        ("Cell-Level", cell_level.get("count", 0),
         cell_level.get("soft_avg", 0), cell_level.get("hard_avg", 0)),
        ("Sheet-Level", sheet_level.get("count", 0),
         sheet_level.get("soft_avg", 0), sheet_level.get("hard_avg", 0)),
    ]

    for i, (cat, count, soft, hard) in enumerate(score_rows, start=row + 1):
        ws_summary.cell(row=i, column=1, value=cat)
        ws_summary.cell(row=i, column=2, value=count)
        ws_summary.cell(row=i, column=3, value=soft).number_format = '0.0%'
        ws_summary.cell(row=i, column=4, value=hard).number_format = '0.0%'

    # Cost section
    row = 19
    ws_summary[f"A{row}"] = "Cost"
    ws_summary[f"A{row}"].font = Font(size=14, bold=True)
    cost_rows = [
        ("Input Tokens", f"{cost.get('input_tokens', 0):,}"),
        ("Output Tokens", f"{cost.get('output_tokens', 0):,}"),
        ("API Calls", cost.get("api_calls", 0)),
        ("Code Executions", cost.get("code_executions", 0)),
        ("Estimated Cost (USD)", f"${cost.get('estimated_usd', 0):.2f}"),
    ]
    for i, (label, value) in enumerate(cost_rows, start=row + 1):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = Font(bold=True)
        ws_summary[f"B{i}"] = value

    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40

    # --- Per-instruction results sheet ---
    ws_results = wb.create_sheet("Results")

    result_headers = [
        "ID", "Instruction Type", "Status",
        "TC1", "TC2", "TC3",
        "Soft Score", "Hard Score",
        "Duration (s)", "Answer Position",
        "Instruction (truncated)",
    ]

    for col, header in enumerate(result_headers, 1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    results_list = results_data.get("results", [])
    for row_idx, r in enumerate(results_list, start=2):
        tc = r.get("test_case_results", [0, 0, 0])
        ws_results.cell(row=row_idx, column=1, value=r.get("id", ""))
        ws_results.cell(row=row_idx, column=2, value=r.get("instruction_type", ""))
        ws_results.cell(row=row_idx, column=3, value=r.get("status", ""))
        ws_results.cell(row=row_idx, column=4, value=tc[0] if len(tc) > 0 else 0)
        ws_results.cell(row=row_idx, column=5, value=tc[1] if len(tc) > 1 else 0)
        ws_results.cell(row=row_idx, column=6, value=tc[2] if len(tc) > 2 else 0)
        ws_results.cell(row=row_idx, column=7, value=r.get("soft_restriction", 0))
        ws_results.cell(row=row_idx, column=7).number_format = '0.00%'
        ws_results.cell(row=row_idx, column=8, value=r.get("hard_restriction", 0))
        ws_results.cell(row=row_idx, column=9, value=round(r.get("duration_sec", 0), 1))
        ws_results.cell(row=row_idx, column=10, value=r.get("answer_position", ""))
        instruction = r.get("instruction", "")
        ws_results.cell(row=row_idx, column=11, value=instruction[:200] if instruction else "")

        # Color-code pass/fail
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for tc_col in [4, 5, 6]:
            cell = ws_results.cell(row=row_idx, column=tc_col)
            cell.fill = green if cell.value == 1 else red

    # Auto-filter
    ws_results.auto_filter.ref = f"A1:K{len(results_list) + 1}"

    # Column widths
    widths = [12, 25, 12, 6, 6, 6, 12, 12, 12, 20, 60]
    for i, w in enumerate(widths, 1):
        ws_results.column_dimensions[get_column_letter(i)].width = w

    # --- Competitive Comparison sheet ---
    ws_comp = wb.create_sheet("Competitive")

    comp_headers = ["Model / Setting", "Soft Score", "Hard Score", "Source"]
    for col, header in enumerate(comp_headers, 1):
        cell = ws_comp.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Our result first (highlighted)
    our_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    our_label = f"ServalSheets ({summary.get('setting', '')}) — {summary.get('model', '')}"
    row = 2
    ws_comp.cell(row=row, column=1, value=our_label).fill = our_fill
    ws_comp.cell(row=row, column=2, value=overall.get("soft_avg", 0)).fill = our_fill
    ws_comp.cell(row=row, column=2).number_format = '0.0%'
    ws_comp.cell(row=row, column=3, value=overall.get("hard_avg", 0)).fill = our_fill
    ws_comp.cell(row=row, column=3).number_format = '0.0%'
    ws_comp.cell(row=row, column=4, value="This run").fill = our_fill

    # Competitors
    for name, scores in sorted(COMPETITOR_SCORES.items(), key=lambda x: -x[1]["hard"]):
        row += 1
        ws_comp.cell(row=row, column=1, value=name)
        ws_comp.cell(row=row, column=2, value=scores["soft"])
        ws_comp.cell(row=row, column=2).number_format = '0.0%'
        ws_comp.cell(row=row, column=3, value=scores["hard"])
        ws_comp.cell(row=row, column=3).number_format = '0.0%'
        ws_comp.cell(row=row, column=4, value="NeurIPS 2024 paper")

    ws_comp.column_dimensions['A'].width = 50
    ws_comp.column_dimensions['B'].width = 15
    ws_comp.column_dimensions['C'].width = 15
    ws_comp.column_dimensions['D'].width = 25

    wb.save(output_path)
    print(f"XLSX report saved to: {output_path}")


# ============================================================================
# HTML Dashboard Generator
# ============================================================================

def generate_html_dashboard(results_data, output_path, all_runs=None):
    """Generate an interactive HTML dashboard with Chart.js."""

    summary = results_data.get("summary", {})
    scores = summary.get("scores", {})
    overall = scores.get("overall", {})
    results_list = results_data.get("results", [])
    cost = summary.get("cost", {})

    # Compute per-status counts
    status_counts = {}
    for r in results_list:
        s = r.get("status", "unknown")
        status_counts[s] = status_counts.get(s, 0) + 1

    # Compute pass/fail distribution for hard scores
    hard_pass = sum(1 for r in results_list if r.get("hard_restriction") == 1)
    hard_fail = len(results_list) - hard_pass

    # Competitive data for chart
    comp_labels = [f"ServalSheets ({summary.get('model', '')})"]
    comp_soft = [overall.get("soft_avg", 0)]
    comp_hard = [overall.get("hard_avg", 0)]
    for name, sc in sorted(COMPETITOR_SCORES.items(), key=lambda x: -x[1]["hard"]):
        comp_labels.append(name)
        comp_soft.append(sc["soft"])
        comp_hard.append(sc["hard"])

    # Error analysis
    error_types = {}
    for r in results_list:
        if r.get("hard_restriction") == 0:
            details = r.get("details", [])
            for d in details:
                if not d.get("passed"):
                    msg = d.get("message", "unknown")
                    # Categorize
                    if "worksheet not found" in msg:
                        cat = "Wrong sheet name"
                    elif "File not exist" in msg:
                        cat = "No output file"
                    elif "Value difference" in msg:
                        cat = "Value mismatch"
                    else:
                        cat = "Other error"
                    error_types[cat] = error_types.get(cat, 0) + 1
                    break

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SpreadsheetBench Results — ServalSheets</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                background: #f5f7fa; color: #333; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}
        h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 8px; color: #1a1a2e; }}
        .subtitle {{ color: #666; margin-bottom: 24px; font-size: 14px; }}
        .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 32px; }}
        .card {{ background: white; border-radius: 12px; padding: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        .card-label {{ font-size: 13px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }}
        .card-value {{ font-size: 32px; font-weight: 700; }}
        .card-value.green {{ color: #10b981; }}
        .card-value.blue {{ color: #3b82f6; }}
        .card-value.amber {{ color: #f59e0b; }}
        .card-value.red {{ color: #ef4444; }}
        .charts {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 32px; }}
        .chart-card {{ background: white; border-radius: 12px; padding: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        .chart-card h3 {{ font-size: 16px; margin-bottom: 16px; color: #1a1a2e; }}
        .chart-card.full {{ grid-column: 1 / -1; }}
        canvas {{ max-height: 400px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #1f4e79; color: white; padding: 10px 12px; text-align: left; font-size: 13px; }}
        td {{ padding: 8px 12px; border-bottom: 1px solid #eee; font-size: 13px; }}
        tr:hover {{ background: #f0f7ff; }}
        .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 12px; font-weight: 600; }}
        .badge.pass {{ background: #d1fae5; color: #065f46; }}
        .badge.fail {{ background: #fee2e2; color: #991b1b; }}
        .badge.ours {{ background: #dbeafe; color: #1e40af; }}
        @media (max-width: 768px) {{ .charts {{ grid-template-columns: 1fr; }} }}
    </style>
</head>
<body>
<div class="container">
    <h1>SpreadsheetBench Results</h1>
    <div class="subtitle">
        {summary.get('track', '')} &middot; {summary.get('setting', '')} &middot;
        {summary.get('model', '')} &middot; {summary.get('timestamp', '')[:10]}
    </div>

    <!-- Summary Cards -->
    <div class="cards">
        <div class="card">
            <div class="card-label">Hard Score</div>
            <div class="card-value blue">{overall.get('hard_pct', '0%')}</div>
        </div>
        <div class="card">
            <div class="card-label">Soft Score</div>
            <div class="card-value green">{overall.get('soft_pct', '0%')}</div>
        </div>
        <div class="card">
            <div class="card-label">Instructions</div>
            <div class="card-value">{summary.get('dataset_size', 0)}</div>
        </div>
        <div class="card">
            <div class="card-label">Completed</div>
            <div class="card-value green">{summary.get('completed', 0)}</div>
        </div>
        <div class="card">
            <div class="card-label">Errors</div>
            <div class="card-value {'red' if summary.get('errors', 0) > 0 else 'green'}">{summary.get('errors', 0)}</div>
        </div>
        <div class="card">
            <div class="card-label">Est. Cost</div>
            <div class="card-value amber">${cost.get('estimated_usd', 0):.2f}</div>
        </div>
    </div>

    <!-- Charts -->
    <div class="charts">
        <!-- Competitive Comparison -->
        <div class="chart-card full">
            <h3>Competitive Comparison (Hard Score %)</h3>
            <canvas id="compChart"></canvas>
        </div>

        <!-- By Instruction Type -->
        <div class="chart-card">
            <h3>Scores by Instruction Type</h3>
            <canvas id="typeChart"></canvas>
        </div>

        <!-- Error Distribution -->
        <div class="chart-card">
            <h3>Error Distribution</h3>
            <canvas id="errorChart"></canvas>
        </div>
    </div>

    <!-- Competitive Table -->
    <div class="chart-card full" style="margin-bottom: 32px;">
        <h3>Detailed Competitive Comparison</h3>
        <table>
            <thead>
                <tr><th>Model / Setting</th><th>Soft Score</th><th>Hard Score</th><th>Source</th></tr>
            </thead>
            <tbody>
                <tr style="background: #eff6ff;">
                    <td><span class="badge ours">OURS</span> ServalSheets ({summary.get('setting', '')}) — {summary.get('model', '')}</td>
                    <td><strong>{overall.get('soft_avg', 0)*100:.1f}%</strong></td>
                    <td><strong>{overall.get('hard_avg', 0)*100:.1f}%</strong></td>
                    <td>This run</td>
                </tr>
"""

    for name, sc in sorted(COMPETITOR_SCORES.items(), key=lambda x: -x[1]["hard"]):
        html += f"""                <tr>
                    <td>{name}</td>
                    <td>{sc['soft']*100:.1f}%</td>
                    <td>{sc['hard']*100:.1f}%</td>
                    <td>NeurIPS 2024 paper</td>
                </tr>
"""

    html += f"""            </tbody>
        </table>
    </div>
</div>

<script>
    // Competitive comparison chart
    new Chart(document.getElementById('compChart'), {{
        type: 'bar',
        data: {{
            labels: {json.dumps(comp_labels)},
            datasets: [
                {{ label: 'Hard Score', data: {json.dumps([round(x*100, 1) for x in comp_hard])},
                   backgroundColor: comp_labels.map((_, i) => i === 0 ? '#3b82f6' : '#94a3b8') }},
                {{ label: 'Soft Score', data: {json.dumps([round(x*100, 1) for x in comp_soft])},
                   backgroundColor: comp_labels.map((_, i) => i === 0 ? '#60a5fa' : '#cbd5e1') }},
            ]
        }},
        options: {{
            indexAxis: 'y',
            responsive: true,
            plugins: {{ legend: {{ position: 'bottom' }} }},
            scales: {{ x: {{ beginAtZero: true, max: 100, title: {{ display: true, text: 'Score (%)' }} }} }}
        }}
    }});

    // By instruction type
    new Chart(document.getElementById('typeChart'), {{
        type: 'bar',
        data: {{
            labels: ['Cell-Level', 'Sheet-Level'],
            datasets: [
                {{ label: 'Soft Score', data: [{scores.get('cell_level', {}).get('soft_avg', 0)*100:.1f}, {scores.get('sheet_level', {}).get('soft_avg', 0)*100:.1f}], backgroundColor: '#10b981' }},
                {{ label: 'Hard Score', data: [{scores.get('cell_level', {}).get('hard_avg', 0)*100:.1f}, {scores.get('sheet_level', {}).get('hard_avg', 0)*100:.1f}], backgroundColor: '#3b82f6' }},
            ]
        }},
        options: {{ responsive: true, scales: {{ y: {{ beginAtZero: true, max: 100 }} }} }}
    }});

    // Error distribution
    new Chart(document.getElementById('errorChart'), {{
        type: 'doughnut',
        data: {{
            labels: {json.dumps(list(error_types.keys()) if error_types else ['Pass', 'Fail'])},
            datasets: [{{ data: {json.dumps(list(error_types.values()) if error_types else [hard_pass, hard_fail])},
                          backgroundColor: ['#ef4444', '#f59e0b', '#8b5cf6', '#10b981'] }}]
        }},
        options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom' }} }} }}
    }});
</script>
</body>
</html>"""

    with open(output_path, 'w') as f:
        f.write(html)

    print(f"HTML dashboard saved to: {output_path}")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Generate reports from SpreadsheetBench results")
    parser.add_argument("--results-dir", required=True, help="Path to results directory")
    parser.add_argument("--output-dir", default=None, help="Output directory (defaults to results-dir)")
    args = parser.parse_args()

    results_dir = Path(args.results_dir)
    output_dir = Path(args.output_dir) if args.output_dir else results_dir

    # Load results
    results_file = results_dir / "results.json"
    if not results_file.exists():
        print(f"ERROR: results.json not found in {results_dir}")
        sys.exit(1)

    with open(results_file) as f:
        results_data = json.load(f)

    # Generate reports
    generate_xlsx_report(results_data, output_dir / "report.xlsx")
    generate_html_dashboard(results_data, output_dir / "dashboard.html")

    print(f"\nAll reports generated in: {output_dir}")


if __name__ == "__main__":
    main()
