"""
Track B Local Executor — ServalSheets MCP tool simulation via openpyxl.

Claude sees the REAL ServalSheets tool schemas and makes tool_use calls.
Tool execution is backed by openpyxl locally instead of Google Sheets API.
This tests the same intelligence: can Claude use ServalSheets tools to solve tasks?

Comparison uses the official SpreadsheetBench evaluation logic (byte-identical).
"""

import os
import sys
import json
import copy
import time
import traceback
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Anthropic client
_anthropic_client = None


def _get_client(api_key):
    global _anthropic_client
    if _anthropic_client is None:
        import httpx
        from anthropic import Anthropic
        # Use the HTTP proxy with SSL verification disabled (proxy uses self-signed cert)
        http_proxy = os.environ.get("HTTP_PROXY", os.environ.get("http_proxy", ""))
        if http_proxy and not http_proxy.startswith("socks"):
            transport = httpx.HTTPTransport(proxy=http_proxy, verify=False)
            http_client = httpx.Client(transport=transport, timeout=180)
            _anthropic_client = Anthropic(api_key=api_key, http_client=http_client)
        else:
            _anthropic_client = Anthropic(api_key=api_key)
    return _anthropic_client


# ============================================================================
# Local Spreadsheet Store — openpyxl backed
# ============================================================================

class LocalSpreadsheetStore:
    """Manages in-memory spreadsheets backed by openpyxl workbooks."""

    def __init__(self):
        self._workbooks = {}  # spreadsheet_id -> openpyxl.Workbook
        self._counter = 0

    def import_xlsx(self, file_path, title=None):
        """Import an XLSX file and return a spreadsheet ID."""
        self._counter += 1
        ssid = f"local_{self._counter}"
        wb = openpyxl.load_workbook(file_path, data_only=False)
        self._workbooks[ssid] = wb
        return ssid

    def export_xlsx(self, ssid, output_path):
        """Export a spreadsheet to an XLSX file."""
        wb = self._workbooks.get(ssid)
        if not wb:
            raise ValueError(f"Spreadsheet {ssid} not found")
        wb.save(output_path)

    def delete(self, ssid):
        """Remove a spreadsheet from memory."""
        self._workbooks.pop(ssid, None)

    def get_workbook(self, ssid):
        return self._workbooks.get(ssid)

    def get_sheet_names(self, ssid):
        wb = self._workbooks.get(ssid)
        return wb.sheetnames if wb else []

    def get_sheet(self, ssid, sheet_name=None):
        wb = self._workbooks.get(ssid)
        if not wb:
            return None
        if sheet_name:
            return wb[sheet_name] if sheet_name in wb.sheetnames else None
        return wb.active


# ============================================================================
# Range Parser
# ============================================================================

def parse_range(range_str):
    """
    Parse A1 notation like "Sheet1!A1:D10" or "A1:D10" or "'My Sheet'!B2:C5".
    Returns (sheet_name, start_row, start_col, end_row, end_col).
    """
    sheet_name = None
    cell_range = range_str

    # Handle sheet name prefix
    if '!' in range_str:
        sheet_part, cell_range = range_str.rsplit('!', 1)
        sheet_name = sheet_part.strip("'\"")

    # Parse cell range
    if ':' in cell_range:
        start, end = cell_range.split(':')
    else:
        start = end = cell_range

    start_col, start_row = _parse_cell_ref(start)
    end_col, end_row = _parse_cell_ref(end)

    return sheet_name, start_row, start_col, end_row, end_col


def _parse_cell_ref(ref):
    """Parse 'A1' into (col_num, row_num)."""
    match = re.match(r'^([A-Za-z]+)(\d+)$', ref.strip())
    if not match:
        return 1, 1
    col_str = match.group(1).upper()
    row = int(match.group(2))
    col = column_index_from_string(col_str)
    return col, row


# ============================================================================
# Tool Call Simulator
# ============================================================================

class ToolSimulator:
    """
    Simulates ServalSheets MCP tool calls using openpyxl.
    Handles the subset of tools/actions needed for benchmark tasks.
    """

    def __init__(self, store: LocalSpreadsheetStore):
        self.store = store

    def execute(self, tool_name, arguments):
        """Execute a tool call and return the result as a dict."""
        if not isinstance(arguments, dict):
            return {"response": {"success": False, "error": f"Invalid arguments type: {type(arguments).__name__}"}}
        request = arguments.get("request", arguments)
        if not isinstance(request, dict):
            return {"response": {"success": False, "error": f"Invalid request type: {type(request).__name__}"}}
        action = request.get("action", "")
        ssid = request.get("spreadsheetId", "")

        try:
            if tool_name == "sheets_core":
                return self._handle_core(action, request, ssid)
            elif tool_name == "sheets_data":
                return self._handle_data(action, request, ssid)
            elif tool_name == "sheets_composite":
                return self._handle_composite(action, request, ssid)
            elif tool_name == "sheets_format":
                return self._handle_format(action, request, ssid)
            elif tool_name == "sheets_dimensions":
                return self._handle_dimensions(action, request, ssid)
            elif tool_name == "sheets_compute":
                return self._handle_compute(action, request, ssid)
            else:
                return {"response": {"success": True, "action": action,
                                     "message": f"Tool {tool_name}.{action} acknowledged (no-op in simulation)"}}
        except Exception as e:
            return {"response": {"success": False, "error": {"code": "SIMULATION_ERROR",
                                                             "message": str(e)}}}

    def _handle_core(self, action, req, ssid):
        if action == "get":
            sheets = self.store.get_sheet_names(ssid)
            wb = self.store.get_workbook(ssid)
            sheet_info = []
            for name in sheets:
                ws = wb[name]
                sheet_info.append({
                    "title": name,
                    "sheetId": sheets.index(name),
                    "rowCount": ws.max_row or 0,
                    "columnCount": ws.max_column or 0,
                })
            return {"response": {"success": True, "action": "get",
                                 "spreadsheetId": ssid, "title": "Benchmark Spreadsheet",
                                 "sheets": sheet_info}}

        elif action == "list_sheets":
            sheets = self.store.get_sheet_names(ssid)
            return {"response": {"success": True, "action": "list_sheets",
                                 "sheets": [{"title": s, "sheetId": i}
                                           for i, s in enumerate(sheets)]}}

        elif action == "add_sheet":
            wb = self.store.get_workbook(ssid)
            name = req.get("sheetName", req.get("title", "NewSheet"))
            wb.create_sheet(title=name)
            return {"response": {"success": True, "action": "add_sheet",
                                 "sheetName": name}}

        elif action == "delete_sheet":
            wb = self.store.get_workbook(ssid)
            name = req.get("sheetName")
            if name and name in wb.sheetnames:
                del wb[name]
            return {"response": {"success": True, "action": "delete_sheet"}}

        elif action == "duplicate_sheet":
            wb = self.store.get_workbook(ssid)
            source = req.get("sheetName") or wb.active.title
            new_name = req.get("newTitle", f"{source}_copy")
            if source in wb.sheetnames:
                src_ws = wb[source]
                new_ws = wb.copy_worksheet(src_ws)
                new_ws.title = new_name
            return {"response": {"success": True, "action": "duplicate_sheet",
                                 "newSheetName": new_name}}

        return {"response": {"success": True, "action": action,
                             "message": f"core.{action} acknowledged"}}

    def _handle_data(self, action, req, ssid):
        if action == "read":
            range_str = req.get("range", "A1:Z100")
            if isinstance(range_str, dict):
                range_str = range_str.get("range", "A1:Z100")
            return self._read_range(ssid, str(range_str),
                                    req.get("valueRenderOption", "FORMATTED_VALUE"))

        elif action == "write":
            range_str = req.get("range", "A1")
            if isinstance(range_str, dict):
                range_str = range_str.get("range", "A1")
            values = req.get("values", [])
            return self._write_range(ssid, str(range_str), values,
                                     req.get("valueInputOption", "USER_ENTERED"))

        elif action == "append":
            range_str = req.get("range", "A1")
            if isinstance(range_str, dict):
                range_str = range_str.get("range", "A1")
            values = req.get("values", [])
            return self._append_range(ssid, str(range_str), values)

        elif action == "clear":
            range_str = req.get("range", "A1:Z1000")
            if isinstance(range_str, dict):
                range_str = range_str.get("range", "A1:Z1000")
            return self._clear_range(ssid, str(range_str))

        elif action == "batch_read":
            ranges = req.get("ranges", [])
            results = []
            for r in ranges:
                r_str = r if isinstance(r, str) else r.get("range", "A1:Z100")
                result = self._read_range(ssid, r_str)
                results.append(result.get("response", {}))
            return {"response": {"success": True, "action": "batch_read",
                                 "results": results}}

        elif action == "batch_write":
            data_items = req.get("data", [])
            written = 0
            for item in data_items:
                r_str = item.get("range", "A1")
                if isinstance(r_str, dict):
                    r_str = r_str.get("range", "A1")
                vals = item.get("values", [])
                self._write_range(ssid, str(r_str), vals)
                written += len(vals)
            return {"response": {"success": True, "action": "batch_write",
                                 "updatedRanges": len(data_items), "updatedRows": written}}

        elif action == "find_replace":
            find_str = req.get("find", "")
            replace_str = req.get("replacement", "")
            return self._find_replace(ssid, find_str, replace_str, req)

        elif action == "copy_paste":
            return {"response": {"success": True, "action": "copy_paste",
                                 "message": "copy_paste acknowledged"}}

        return {"response": {"success": True, "action": action,
                             "message": f"data.{action} acknowledged"}}

    def _handle_composite(self, action, req, ssid):
        if action == "import_xlsx":
            # Already handled at higher level
            return {"response": {"success": True, "action": "import_xlsx"}}
        elif action == "export_xlsx":
            return {"response": {"success": True, "action": "export_xlsx"}}
        return {"response": {"success": True, "action": action,
                             "message": f"composite.{action} acknowledged"}}

    def _handle_format(self, action, req, ssid):
        # Format operations don't affect cell values, so they're no-ops for benchmark
        return {"response": {"success": True, "action": action,
                             "message": f"format.{action} applied (visual only)"}}

    def _handle_dimensions(self, action, req, ssid):
        if action == "sort_range":
            return self._sort_range(ssid, req)
        elif action == "delete_rows":
            return self._delete_rows(ssid, req)
        elif action == "delete_columns":
            return self._delete_columns(ssid, req)
        elif action == "insert_rows":
            return self._insert_rows(ssid, req)
        elif action == "insert_columns":
            return self._insert_columns(ssid, req)
        return {"response": {"success": True, "action": action,
                             "message": f"dimensions.{action} acknowledged"}}

    def _handle_compute(self, action, req, ssid):
        # Compute actions return statistical results — simplified simulation
        return {"response": {"success": True, "action": action,
                             "message": f"compute.{action} — use sheets_data.read to get raw data and compute in your response"}}

    # --- Data Operations ---

    def _read_range(self, ssid, range_str, render_option="FORMATTED_VALUE"):
        sheet_name, sr, sc, er, ec = parse_range(range_str)
        ws = self.store.get_sheet(ssid, sheet_name)
        if not ws:
            sheet_names = self.store.get_sheet_names(ssid)
            return {"response": {"success": False, "error": {
                "code": "SHEET_NOT_FOUND",
                "message": f"Sheet '{sheet_name}' not found. Available: {sheet_names}"}}}

        values = []
        for row in range(sr, er + 1):
            row_data = []
            for col in range(sc, ec + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None:
                    row_data.append("")
                elif isinstance(val, (int, float)):
                    row_data.append(val)
                else:
                    row_data.append(str(val))
            values.append(row_data)

        return {"response": {"success": True, "action": "read",
                             "range": range_str, "values": values,
                             "rowCount": len(values),
                             "columnCount": len(values[0]) if values else 0}}

    def _write_range(self, ssid, range_str, values, input_option="USER_ENTERED"):
        sheet_name, sr, sc, er, ec = parse_range(range_str)
        ws = self.store.get_sheet(ssid, sheet_name)
        if not ws:
            # Create sheet if it doesn't exist
            wb = self.store.get_workbook(ssid)
            if wb and sheet_name:
                wb.create_sheet(title=sheet_name)
                ws = wb[sheet_name]
            else:
                ws = self.store.get_sheet(ssid)

        cells_written = 0
        for i, row_data in enumerate(values):
            if not isinstance(row_data, (list, tuple)):
                row_data = [row_data]
            for j, val in enumerate(row_data):
                r = sr + i
                c = sc + j
                cell = ws.cell(row=r, column=c)

                # Handle value types
                if val is None or val == "":
                    cell.value = None
                elif isinstance(val, str) and val.startswith("="):
                    # Formula — store as value since openpyxl can't evaluate
                    cell.value = val
                elif isinstance(val, str) and input_option == "USER_ENTERED":
                    # Try numeric parse
                    try:
                        if '.' in val:
                            cell.value = float(val)
                        else:
                            cell.value = int(val)
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = val
                cells_written += 1

        return {"response": {"success": True, "action": "write",
                             "updatedRange": range_str, "updatedCells": cells_written}}

    def _append_range(self, ssid, range_str, values):
        sheet_name, sr, sc, er, ec = parse_range(range_str)
        ws = self.store.get_sheet(ssid, sheet_name)
        if not ws:
            return {"response": {"success": False, "error": {"message": "Sheet not found"}}}

        # Find last row with data
        last_row = ws.max_row or 1
        next_row = last_row + 1

        for i, row_data in enumerate(values):
            if not isinstance(row_data, (list, tuple)):
                row_data = [row_data]
            for j, val in enumerate(row_data):
                ws.cell(row=next_row + i, column=sc + j, value=val)

        return {"response": {"success": True, "action": "append",
                             "updatedRange": f"{sheet_name or 'Sheet1'}!{get_column_letter(sc)}{next_row}",
                             "updatedRows": len(values)}}

    def _clear_range(self, ssid, range_str):
        sheet_name, sr, sc, er, ec = parse_range(range_str)
        ws = self.store.get_sheet(ssid, sheet_name)
        if not ws:
            return {"response": {"success": True, "action": "clear", "clearedCells": 0}}

        cleared = 0
        for row in range(sr, er + 1):
            for col in range(sc, ec + 1):
                ws.cell(row=row, column=col).value = None
                cleared += 1

        return {"response": {"success": True, "action": "clear", "clearedCells": cleared}}

    def _find_replace(self, ssid, find_str, replace_str, req):
        wb = self.store.get_workbook(ssid)
        if not wb:
            return {"response": {"success": False, "error": {"message": "Spreadsheet not found"}}}

        count = 0
        match_case = req.get("matchCase", False)
        all_sheets = req.get("allSheets", True)
        sheets_to_search = wb.sheetnames if all_sheets else [wb.active.title]

        for sn in sheets_to_search:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if match_case:
                            if find_str in cell.value:
                                cell.value = cell.value.replace(find_str, replace_str)
                                count += 1
                        else:
                            if find_str.lower() in cell.value.lower():
                                cell.value = re.sub(re.escape(find_str), replace_str,
                                                   cell.value, flags=re.IGNORECASE)
                                count += 1

        return {"response": {"success": True, "action": "find_replace",
                             "replacementsCount": count}}

    def _sort_range(self, ssid, req):
        range_str = req.get("range", "A1:Z1000")
        if isinstance(range_str, dict):
            range_str = range_str.get("range", "A1:Z1000")
        sheet_name, sr, sc, er, ec = parse_range(str(range_str))
        ws = self.store.get_sheet(ssid, sheet_name)
        if not ws:
            return {"response": {"success": True, "action": "sort_range"}}

        sort_col = req.get("sortColumn", req.get("column", 0))
        ascending = req.get("ascending", req.get("order", "ASCENDING")) != "DESCENDING"

        # Read all data in range
        data = []
        for row in range(sr, er + 1):
            row_data = []
            for col in range(sc, ec + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)

        # Skip header if present
        has_header = req.get("hasHeader", True)
        header = data[0] if has_header and data else None
        sort_data = data[1:] if has_header else data

        # Sort
        col_idx = sort_col if isinstance(sort_col, int) else 0
        try:
            sort_data.sort(key=lambda r: (r[col_idx] is None, r[col_idx] if r[col_idx] is not None else ""),
                          reverse=not ascending)
        except (IndexError, TypeError):
            pass

        if header:
            sort_data = [header] + sort_data

        # Write back
        for i, row_data in enumerate(sort_data):
            for j, val in enumerate(row_data):
                ws.cell(row=sr + i, column=sc + j, value=val)

        return {"response": {"success": True, "action": "sort_range", "sortedRows": len(sort_data)}}

    def _delete_rows(self, ssid, req):
        ws = self.store.get_sheet(ssid, req.get("sheetName"))
        if ws:
            start = req.get("startRow", req.get("startIndex", 1))
            count = req.get("count", req.get("numRows", 1))
            ws.delete_rows(start, count)
        return {"response": {"success": True, "action": "delete_rows"}}

    def _delete_columns(self, ssid, req):
        ws = self.store.get_sheet(ssid, req.get("sheetName"))
        if ws:
            start = req.get("startColumn", req.get("startIndex", 1))
            count = req.get("count", req.get("numColumns", 1))
            ws.delete_cols(start, count)
        return {"response": {"success": True, "action": "delete_columns"}}

    def _insert_rows(self, ssid, req):
        ws = self.store.get_sheet(ssid, req.get("sheetName"))
        if ws:
            start = req.get("startRow", req.get("startIndex", 1))
            count = req.get("count", req.get("numRows", 1))
            ws.insert_rows(start, count)
        return {"response": {"success": True, "action": "insert_rows"}}

    def _insert_columns(self, ssid, req):
        ws = self.store.get_sheet(ssid, req.get("sheetName"))
        if ws:
            start = req.get("startColumn", req.get("startIndex", 1))
            count = req.get("count", req.get("numColumns", 1))
            ws.insert_cols(start, count)
        return {"response": {"success": True, "action": "insert_columns"}}


# ============================================================================
# Claude Tool Definitions (matches real ServalSheets MCP schemas)
# ============================================================================

TOOL_DEFINITIONS = [
    {
        "name": "sheets_core",
        "description": "Spreadsheet/sheet management. Actions: get, list_sheets, add_sheet, delete_sheet, duplicate_sheet, clear_sheet, update_sheet. Request format: {\"request\": {\"action\": \"...\", \"spreadsheetId\": \"...\", ...}}",
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string", "enum": ["get", "list_sheets", "add_sheet", "delete_sheet", "duplicate_sheet", "clear_sheet", "update_sheet"]},
                        "spreadsheetId": {"type": "string"},
                        "sheetName": {"type": "string"},
                        "sheetId": {"type": "number"},
                        "newTitle": {"type": "string"},
                        "title": {"type": "string"},
                    },
                    "required": ["action"]
                }
            },
            "required": ["request"]
        }
    },
    {
        "name": "sheets_data",
        "description": "Read/write cell values. Actions: read, write, append, clear, batch_read, batch_write, find_replace. Range format: \"Sheet1!A1:D10\". Request format: {\"request\": {\"action\": \"...\", \"spreadsheetId\": \"...\", \"range\": \"...\", ...}}. For write: values is array of arrays [[row1col1, row1col2], [row2col1, row2col2]]. valueInputOption: USER_ENTERED (parse formulas/numbers) or RAW.",
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string", "enum": ["read", "write", "append", "clear", "batch_read", "batch_write", "find_replace", "copy_paste"]},
                        "spreadsheetId": {"type": "string"},
                        "range": {},
                        "ranges": {"type": "array"},
                        "values": {"type": "array"},
                        "data": {"type": "array"},
                        "find": {"type": "string"},
                        "replacement": {"type": "string"},
                        "matchCase": {"type": "boolean"},
                        "allSheets": {"type": "boolean"},
                        "valueInputOption": {"type": "string", "enum": ["RAW", "USER_ENTERED"]},
                        "valueRenderOption": {"type": "string", "enum": ["FORMATTED_VALUE", "UNFORMATTED_VALUE", "FORMULA"]},
                    },
                    "required": ["action"]
                }
            },
            "required": ["request"]
        }
    },
    {
        "name": "sheets_dimensions",
        "description": "Row/column operations. Actions: sort_range, delete_rows, delete_columns, insert_rows, insert_columns, auto_resize, freeze, hide, unhide. Request format: {\"request\": {\"action\": \"...\", \"spreadsheetId\": \"...\", ...}}",
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string", "enum": ["sort_range", "delete_rows", "delete_columns", "insert_rows", "insert_columns", "auto_resize", "freeze", "hide", "unhide", "move_rows", "move_columns"]},
                        "spreadsheetId": {"type": "string"},
                        "sheetName": {"type": "string"},
                        "range": {},
                        "startRow": {"type": "number"},
                        "startColumn": {"type": "number"},
                        "count": {"type": "number"},
                        "numRows": {"type": "number"},
                        "numColumns": {"type": "number"},
                        "startIndex": {"type": "number"},
                        "sortColumn": {"type": "number"},
                        "ascending": {"type": "boolean"},
                        "order": {"type": "string"},
                        "hasHeader": {"type": "boolean"},
                        "column": {"type": "number"},
                    },
                    "required": ["action"]
                }
            },
            "required": ["request"]
        }
    },
    {
        "name": "sheets_format",
        "description": "Cell formatting (bold, colors, borders, number formats). Actions: set_bold, set_background_color, set_number_format, add_conditional_format_rule, set_borders, etc. Visual-only, does not change cell values.",
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string"},
                        "spreadsheetId": {"type": "string"},
                        "range": {},
                    },
                    "required": ["action"]
                }
            },
            "required": ["request"]
        }
    },
    {
        "name": "sheets_compute",
        "description": "Statistical operations: descriptive_stats, regression, correlation, frequency, percentile, forecast, moving_average. Use sheets_data.read to get raw data, compute in your response, then write results with sheets_data.write.",
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string"},
                        "spreadsheetId": {"type": "string"},
                        "range": {},
                    },
                    "required": ["action"]
                }
            },
            "required": ["request"]
        }
    },
]


# ============================================================================
# Spreadsheet Content Preview
# ============================================================================

def gen_preview(wb, max_rows=6, max_sheets=5):
    """Generate text preview of workbook content (matches SpreadsheetBench format)."""
    preview = ""
    for sheet_name in wb.sheetnames[:max_sheets]:
        ws = wb[sheet_name]
        preview += f"Sheet Name: {sheet_name}\n"
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows,
                                                    max_col=min(ws.max_column or 1, 26),
                                                    values_only=True)):
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, (int, float)):
                    cells.append(str(v))
                else:
                    cells.append(str(v)[:50])
            preview += "\t".join(cells) + "\n"
        preview += "-" * 50 + "\n"
    return preview


# ============================================================================
# Agentic Loop
# ============================================================================

def run_agentic_loop(api_key, model, instruction, ssid, answer_position,
                     instruction_type, preview, store, simulator, cost_tracker,
                     max_iterations=15):
    """
    Run Claude with ServalSheets tool definitions in an agentic loop.
    Tool calls are executed locally via the ToolSimulator.
    """
    client = _get_client(api_key)

    system_prompt = f"""You manipulate spreadsheets using tools. SPREADSHEET ID: {ssid}

CRITICAL: Do NOT write long explanations. Make tool calls IMMEDIATELY. No planning text.

1. Read data: sheets_data action "read"
2. Write to: {answer_position} (type: {instruction_type})
3. COMPUTE actual values yourself (SUM, AVERAGE, etc.) — NEVER write formulas, write the result.
4. Envelope: {{"request": {{"action": "...", "spreadsheetId": "...", ...}}}}
5. Say "DONE" when finished. Keep ALL text responses under 100 words.

PREVIEW:
{preview}"""

    user_msg = f"""Complete this spreadsheet manipulation:

{instruction}

Spreadsheet ID: {ssid}
Answer position (cells to modify): {answer_position}
Instruction type: {instruction_type}

Read the relevant data first, then compute and write the answer values."""

    messages = [{"role": "user", "content": user_msg}]
    tool_calls_made = []
    consecutive_max_tokens = 0

    for iteration in range(max_iterations):
        try:
            t0 = time.time()
            response = client.messages.create(
                model=model,
                max_tokens=8192,
                system=system_prompt,
                messages=messages,
                tools=TOOL_DEFINITIONS,
            )
            elapsed = time.time() - t0
            print(f"      iter {iteration}: {response.stop_reason} | {response.usage.input_tokens}in/{response.usage.output_tokens}out | {elapsed:.1f}s", flush=True)
        except Exception as e:
            print(f"      LLM error at iteration {iteration}: {e}", flush=True)
            break

        if cost_tracker:
            cost_tracker.record(tokens={
                "input": response.usage.input_tokens,
                "output": response.usage.output_tokens,
            })

        has_tool_use = any(
            getattr(block, 'type', None) == "tool_use"
            for block in response.content
        )

        if response.stop_reason == "max_tokens":
            consecutive_max_tokens += 1
            if consecutive_max_tokens >= 3:
                print(f"      Breaking: {consecutive_max_tokens} consecutive max_tokens", flush=True)
                break
            if not has_tool_use:
                messages.append({"role": "assistant", "content": response.content})
                messages.append({"role": "user", "content": "Truncated. Make a tool call directly — no explanation needed."})
                continue
        else:
            consecutive_max_tokens = 0

        if not has_tool_use:
            break

        messages.append({"role": "assistant", "content": response.content})

        tool_results = []
        for block in response.content:
            if getattr(block, 'type', None) == "tool_use":
                tool_name = block.name
                tool_input = block.input

                try:
                    sim_result = simulator.execute(tool_name, tool_input)
                    result_text = json.dumps(sim_result, default=str)
                except Exception as ex:
                    result_text = json.dumps({"response": {"success": False, "error": str(ex)}})

                tool_calls_made.append({
                    "tool": tool_name,
                    "action": tool_input.get("request", {}).get("action", "?")
                             if isinstance(tool_input, dict) else "?",
                    "iteration": iteration,
                })

                # Truncate large results to reduce context window bloat
                max_result = 6000
                if len(result_text) > max_result:
                    result_text = result_text[:max_result] + "\n... [TRUNCATED — use a smaller range or read specific columns]"
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result_text,
                })

        messages.append({"role": "user", "content": tool_results})

    return tool_calls_made


# ============================================================================
# Instruction Executor
# ============================================================================

class LocalMcpExecutor:
    """Execute SpreadsheetBench instructions using local MCP simulation."""

    def __init__(self, api_key, model="claude-sonnet-4-20250514"):
        self.api_key = api_key
        self.model = model
        self.store = LocalSpreadsheetStore()
        self.simulator = ToolSimulator(self.store)

    def execute_instruction(self, data, dataset_path, cost_tracker, timeout_per_tc=300):
        """Execute all 3 test cases for one instruction."""
        task_id = data['id']
        start_time = time.time()
        test_case_results = []
        test_case_details = []
        all_tool_calls = []
        print(f"    [{task_id}] Starting ({data['instruction_type']})", flush=True)

        for tc_idx in range(3):
            tc_start = time.time()
            input_path = str(
                Path(dataset_path) / "spreadsheet" / str(task_id) /
                f"{tc_idx + 1}_{task_id}_input.xlsx"
            )
            answer_path = str(
                Path(dataset_path) / "spreadsheet" / str(task_id) /
                f"{tc_idx + 1}_{task_id}_answer.xlsx"
            )

            ssid = None
            output_path = None

            try:
                # 1. Import XLSX locally
                ssid = self.store.import_xlsx(input_path, title=f"Bench_{task_id}_TC{tc_idx+1}")

                # 2. Generate preview
                wb = self.store.get_workbook(ssid)
                preview = gen_preview(wb)

                # 3. Run agentic loop
                tool_calls = run_agentic_loop(
                    api_key=self.api_key,
                    model=self.model,
                    instruction=data['instruction'],
                    ssid=ssid,
                    answer_position=data['answer_position'],
                    instruction_type=data['instruction_type'],
                    preview=preview,
                    store=self.store,
                    simulator=self.simulator,
                    cost_tracker=cost_tracker,
                )
                all_tool_calls.extend(tool_calls)

                # 4. Export result
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                    output_path = f.name
                self.store.export_xlsx(ssid, output_path)

                # 5. Compare using official logic
                sys.path.insert(0, str(Path(__file__).parent.parent / "evaluation"))
                from compare import compare_workbooks
                result, msg, mismatches = compare_workbooks(
                    answer_path, output_path,
                    data['instruction_type'],
                    data['answer_position'],
                )

                test_case_results.append(int(result))
                tc_dur = round(time.time() - tc_start, 1)
                test_case_details.append({
                    'test_case': tc_idx + 1,
                    'passed': result,
                    'message': msg,
                    'tool_calls': len(tool_calls),
                    'duration_sec': tc_dur,
                })
                print(f"    [{task_id}] TC{tc_idx+1}: {'PASS' if result else 'FAIL'} ({tc_dur}s, {len(tool_calls)} tools)", flush=True)

            except Exception as e:
                tc_dur = round(time.time() - tc_start, 1)
                test_case_results.append(0)
                test_case_details.append({
                    'test_case': tc_idx + 1,
                    'passed': False,
                    'message': str(e),
                    'traceback': traceback.format_exc(),
                    'duration_sec': tc_dur,
                })
                print(f"    [{task_id}] TC{tc_idx+1}: ERROR ({tc_dur}s) {str(e)[:100]}", flush=True)

            finally:
                if output_path and os.path.exists(output_path):
                    os.unlink(output_path)
                if ssid:
                    self.store.delete(ssid)

        soft = test_case_results.count(1) / len(test_case_results) if test_case_results else 0
        hard = 0 if 0 in test_case_results else 1

        return {
            'id': task_id,
            'status': 'completed',
            'duration_sec': round(time.time() - start_time, 1),
            'test_case_results': test_case_results,
            'soft_restriction': soft,
            'hard_restriction': hard,
            'details': test_case_details,
            'tool_calls_total': len(all_tool_calls),
            'tool_call_breakdown': all_tool_calls,
        }
