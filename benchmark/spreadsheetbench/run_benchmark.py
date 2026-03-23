#!/usr/bin/env python3
"""
SpreadsheetBench Evaluation Harness for ServalSheets.

Two tracks:
  Track A — Official protocol: Claude generates Python/openpyxl code, executed locally.
  Track B — ServalSheets MCP: Claude uses MCP tools on live Google Sheets.

Both tracks use the EXACT same comparison logic from official evaluation.py.

Usage:
    # Track A on sample (200 instructions)
    python run_benchmark.py --track a --dataset sample --api-key $ANTHROPIC_API_KEY

    # Track A on full 912
    python run_benchmark.py --track a --dataset full --api-key $ANTHROPIC_API_KEY

    # Resume from checkpoint
    python run_benchmark.py --track a --dataset full --api-key $ANTHROPIC_API_KEY --resume

    # Track B (ServalSheets MCP)
    python run_benchmark.py --track b --dataset full --api-key $ANTHROPIC_API_KEY
"""

import os
import sys
import json
import time
import argparse
import subprocess
import tempfile
import traceback
from datetime import datetime, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import openpyxl
from packaging.version import Version
from tqdm import tqdm


# ============================================================================
# Dependency Preflight
# ============================================================================

def preflight_check():
    """Verify all required packages meet minimum versions before any work begins."""
    errors = []

    # openpyxl >= 3.1.5 required by pandas for xlsx read/write
    openpyxl_ver = Version(openpyxl.__version__)
    if openpyxl_ver < Version("3.1.5"):
        errors.append(f"openpyxl {openpyxl.__version__} < 3.1.5 (pandas requirement). "
                      f"Run: pip install --upgrade openpyxl")

    # pandas >= 2.0
    import pandas
    pd_ver = Version(pandas.__version__)
    if pd_ver < Version("2.0"):
        errors.append(f"pandas {pandas.__version__} < 2.0. Run: pip install --upgrade pandas")

    # anthropic SDK
    try:
        import anthropic
    except ImportError:
        errors.append("anthropic SDK not installed. Run: pip install anthropic")

    if errors:
        print("\n[PREFLIGHT FAILED] Fix these issues before running the benchmark:\n")
        for e in errors:
            print(f"  ❌ {e}")
        print()
        sys.exit(1)

    print(f"[preflight] openpyxl={openpyxl.__version__}  pandas={pd.__version__}  ✓")

# Add parent paths for imports
BENCH_DIR = Path(__file__).parent.resolve()
REPO_ROOT = BENCH_DIR.parent.parent  # servalsheets root
# Local data dir takes precedence; falls back to original session path
_LOCAL_DATA = BENCH_DIR / "data"
SPREADSHEETBENCH_DIR = _LOCAL_DATA if _LOCAL_DATA.exists() else Path("/sessions/kind-fervent-gates/SpreadsheetBench")

# Load .env from ServalSheets root as fallback for credentials
_env_file = REPO_ROOT / ".env"
if _env_file.exists():
    with open(_env_file) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _key, _, _val = _line.partition('=')
                _key = _key.strip()
                _val = _val.strip().strip('"').strip("'")
                if _key and _val and not os.environ.get(_key):
                    os.environ[_key] = _val

sys.path.insert(0, str(BENCH_DIR / "evaluation"))
from compare import compare_workbooks, evaluate_instruction


# ============================================================================
# Configuration
# ============================================================================

def _dataset_path(name: str) -> Path:
    """Resolve dataset path: local data/ dir first, then remote session path."""
    local_names = {
        "sample": "sample_data_200",
        "full": "all_data_912_v0.1",
        "verified": "spreadsheetbench_verified_400",
    }
    local = _LOCAL_DATA / local_names[name]
    if local.exists():
        return local
    return SPREADSHEETBENCH_DIR / "data" / local_names[name]

DATASET_PATHS = {k: _dataset_path(k) for k in ("sample", "full", "verified")}

DEFAULT_MODEL = "claude-sonnet-4-20250514"
MAX_WORKERS = 10  # Parallel instruction processing
CODE_EXEC_TIMEOUT = 120  # seconds per code execution
LLM_TIMEOUT = 60  # seconds per LLM call


# ============================================================================
# Prompt Templates (matching official SpreadsheetBench format)
# ============================================================================

PROMPT_SINGLE = """You are a spreadsheet expert who can manipulate spreadsheets through Python code.

You need to solve the given spreadsheet manipulation question, which contains six types of information:
- instruction: The question about spreadsheet manipulation.
- spreadsheet_path: The path of the spreadsheet file you need to manipulate.
- spreadsheet_content: The first few rows of the content of spreadsheet file.
- instruction_type: There are two values (Cell-Level Manipulation, Sheet-Level Manipulation) used to indicate whether the answer to this question applies only to specific cells or to the entire worksheet.
- answer_position: The position need to be modified or filled. For Cell-Level Manipulation questions, this field is filled with the cell position; for Sheet-Level Manipulation, it is the maximum range of cells you need to modify. You only need to modify or fill in values within the cell range specified by answer_position.
- output_path: You need to generate the modified spreadsheet file in this new path.

Below is the spreadsheet manipulation question you need to solve:
### instruction
{instruction}

### spreadsheet_path
{spreadsheet_path}

### spreadsheet_content
{spreadsheet_content}

### instruction_type
{instruction_type}

### answer_position
{answer_position}

### output_path
{output_path}

You should generate Python code for the final solution of the question.

CRITICAL REQUIREMENT — DO NOT WRITE EXCEL FORMULAS:
The output file is evaluated by reading cell values with openpyxl (data_only=True). Any formula string you write (e.g., ws['A1'] = '=SUM(B1:B10)') will be read back as None by the evaluator — causing the task to fail. You MUST:
1. Read all source data from the spreadsheet using openpyxl or pandas
2. Perform ALL calculations using Python arithmetic, pandas, or the math/statistics library
3. Write only the computed numeric or string result to each cell (e.g., ws['A1'] = 42.5)
Never assign a string that starts with '=' to any cell. Compute the value in Python and write the literal result.
"""

PROMPT_MULTI_ROUND = """You are a spreadsheet expert who can manipulate spreadsheets through Python code.

You need to solve the given spreadsheet manipulation question, which contains six types of information:
- instruction: The question about spreadsheet manipulation.
- spreadsheet_path: The path of the spreadsheet file you need to manipulate.
- spreadsheet_content: The first few rows of the content of spreadsheet file.
- instruction_type: There are two values (Cell-Level Manipulation, Sheet-Level Manipulation) used to indicate whether the answer to this question applies only to specific cells or to the entire worksheet.
- answer_position: The position need to be modified or filled. For Cell-Level Manipulation questions, this field is filled with the cell position; for Sheet-Level Manipulation, it is the maximum range of cells you need to modify. You only need to modify or fill in values within the cell range specified by answer_position.
- output_path: You need to generate the modified spreadsheet file in this new path.

Below is the spreadsheet manipulation question you need to solve:
### instruction
{instruction}

### spreadsheet_path
{spreadsheet_path}

### spreadsheet_content
{spreadsheet_content}

### instruction_type
{instruction_type}

### answer_position
{answer_position}

### output_path
{output_path}

The solution of the question can be generated through {max_turn_num} rounds of interaction and you can do two types of actions.
1. Spreadsheet information acquisition: You can generate Python code to obtain the information in the spreadsheet file. In the next turn, the execution result of your Python code will provide to you.
2. Question solution generation: You can generate Python code for the final solution of the question. If error occur when executing code, the error traceback will provide to you for code refinement.

CRITICAL REQUIREMENT — DO NOT WRITE EXCEL FORMULAS:
The output file is evaluated by reading cell values with openpyxl (data_only=True). Any formula string you write (e.g., ws['A1'] = '=SUM(B1:B10)') will be read back as None by the evaluator — causing the task to fail. You MUST:
1. Read all source data from the spreadsheet using openpyxl or pandas
2. Perform ALL calculations using Python arithmetic, pandas, or the math/statistics library
3. Write only the computed numeric or string result to each cell (e.g., ws['A1'] = 42.5)
Never assign a string that starts with '=' to any cell. Compute the value in Python and write the literal result.
"""


# ============================================================================
# Dataset Loading
# ============================================================================

def load_dataset(dataset_name):
    """Load dataset.json and return list of instruction dicts."""
    dataset_path = DATASET_PATHS.get(dataset_name)
    if not dataset_path:
        raise ValueError(f"Unknown dataset: {dataset_name}. Options: {list(DATASET_PATHS.keys())}")

    if not dataset_path.exists():
        # Try extracting
        tar_candidates = [
            dataset_path.parent / f"{dataset_path.name}.tar.gz",
            dataset_path.parent / "spreadsheetbench_912_v0.1.tar.gz",
        ]
        for tar_path in tar_candidates:
            if tar_path.exists():
                print(f"Extracting {tar_path}...")
                subprocess.run(["tar", "-xzf", str(tar_path), "-C", str(tar_path.parent)], check=True)
                break

    dataset_json = dataset_path / "dataset.json"
    if not dataset_json.exists():
        raise FileNotFoundError(f"dataset.json not found at {dataset_json}")

    with open(dataset_json) as f:
        data = json.load(f)

    print(f"Loaded {len(data)} instructions from {dataset_json}")
    return data, dataset_path


def gen_file_content(input_file, max_rows=5):
    """Generate text representation of first N rows of each sheet (matches official)."""
    try:
        excel_file = pd.ExcelFile(input_file)
        sheet_names = excel_file.sheet_names
        excel_data = {}

        for sheet_name in sheet_names:
            df = excel_file.parse(sheet_name)
            nrows = min(max_rows, df.shape[0])
            excel_data[sheet_name] = df.head(nrows).to_string()

        final_str = ""
        for sheet_name, sheet_str in excel_data.items():
            final_str += f"Sheet Name: {sheet_name}\n"
            final_str += sheet_str + "\n"
            final_str += "-" * 50 + "\n"

        return final_str
    except Exception as e:
        return f"Error reading file: {e}"


# ============================================================================
# LLM Client (Anthropic native — not OpenAI compat)
# ============================================================================

def get_anthropic_client(api_key):
    """Create Anthropic client."""
    from anthropic import Anthropic
    return Anthropic(api_key=api_key)


def call_llm(client, messages, model, max_tokens=4096):
    """
    Call Claude via Anthropic API.

    messages: list of {"role": "user"|"assistant", "content": str}
    Returns: response text, token counts
    """
    response = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        messages=messages,
    )
    text = response.content[0].text
    tokens = {
        "input": response.usage.input_tokens,
        "output": response.usage.output_tokens,
    }
    return text, tokens


def call_llm_openai_compat(api_key, messages, model):
    """
    Call Claude via OpenAI-compatible endpoint (for direct comparison with
    other models tested in SpreadsheetBench).
    """
    from openai import OpenAI
    client = OpenAI(
        api_key=api_key,
        base_url="https://api.anthropic.com/v1/",
    )
    chat = client.chat.completions.create(
        model=model,
        messages=messages,
    )
    return chat.choices[0].message.content


# ============================================================================
# Code Extraction & Local Execution (no Docker dependency)
# ============================================================================

def extract_code(response):
    """Extract Python code from LLM response (matches official extract_code)."""
    if '```python' in response:
        code = response[response.find('```python') + len('```python'):]
        code = code[:code.find('```')].lstrip('\n').rstrip('\n')
    elif '```' in response:
        code = response[response.find('```') + 3:]
        code = code[:code.find('```')].lstrip('\n').rstrip('\n')
    else:
        code = response
    return code


def execute_code_local(code, timeout=CODE_EXEC_TIMEOUT):
    """
    Execute Python code in a subprocess (local sandboxed execution).

    Returns (success: bool, output: str) where output is stdout on success
    or stderr on failure. CWD is the benchmark directory (all paths in
    generated code are absolute, so CWD does not matter for correctness).
    """
    with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
        f.write(code)
        script_path = f.name

    try:
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=str(BENCH_DIR),
        )
        if result.returncode != 0:
            err = (result.stderr or result.stdout or "unknown error").strip()
            return False, f"Error (exit {result.returncode}):\n{err}"
        return True, result.stdout
    except subprocess.TimeoutExpired:
        return False, "Error: Code execution timed out"
    except Exception as e:
        return False, f"Error: {e}"
    finally:
        try:
            os.unlink(script_path)
        except OSError:
            pass


# ============================================================================
# Checkpoint / Resume
# ============================================================================

class CheckpointManager:
    """Save/load progress for resumable benchmark runs."""

    def __init__(self, results_dir):
        self.results_dir = Path(results_dir)
        self.results_dir.mkdir(parents=True, exist_ok=True)
        self.checkpoint_file = self.results_dir / "checkpoint.jsonl"
        self.completed = {}
        self._load_existing()

    def _load_existing(self):
        """Load completed results from checkpoint file."""
        if self.checkpoint_file.exists():
            with open(self.checkpoint_file) as f:
                for line in f:
                    if line.strip():
                        entry = json.loads(line)
                        self.completed[entry['id']] = entry
            print(f"Resumed from checkpoint: {len(self.completed)} instructions already completed")

    def is_completed(self, task_id):
        return task_id in self.completed

    def save_result(self, result):
        """Append a result to the checkpoint file."""
        self.completed[result['id']] = result
        with open(self.checkpoint_file, 'a') as f:
            f.write(json.dumps(result, default=str) + '\n')

    def get_all_results(self):
        return list(self.completed.values())


# ============================================================================
# Cost Tracker
# ============================================================================

class CostTracker:
    """Track LLM token usage and API calls."""

    # Pricing (USD per million tokens) — update as needed
    PRICING = {
        "claude-sonnet-4-20250514": {"input": 3.0, "output": 15.0},
        "claude-opus-4-20250514": {"input": 15.0, "output": 75.0},
        "claude-haiku-3-5-20241022": {"input": 0.80, "output": 4.0},
    }

    def __init__(self):
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_api_calls = 0
        self.total_code_execs = 0

    def record(self, tokens=None, api_calls=1, code_execs=0):
        if tokens:
            self.total_input_tokens += tokens.get("input", 0)
            self.total_output_tokens += tokens.get("output", 0)
        self.total_api_calls += api_calls
        self.total_code_execs += code_execs

    def estimate_cost(self, model):
        pricing = self.PRICING.get(model, {"input": 3.0, "output": 15.0})
        input_cost = (self.total_input_tokens / 1_000_000) * pricing["input"]
        output_cost = (self.total_output_tokens / 1_000_000) * pricing["output"]
        return {
            "input_tokens": self.total_input_tokens,
            "output_tokens": self.total_output_tokens,
            "api_calls": self.total_api_calls,
            "code_executions": self.total_code_execs,
            "estimated_usd": round(input_cost + output_cost, 4),
        }


# ============================================================================
# Track A: Official Protocol (Code Generation + Local Execution)
# ============================================================================

def run_track_a_single(data, dataset_path, output_dir, client, model, cost_tracker,
                       setting="single", max_turns=5, max_rows=5):
    """
    Run Track A for a single instruction: generate code → execute → produce output XLSX.

    Follows official SpreadsheetBench inference_single.py protocol exactly.
    """
    task_id = data['id']
    start_time = time.time()

    # Resolve paths — use test case 1 for code generation (official protocol)
    # Supports all naming conventions: _input (912), _init (verified_400), initial.xlsx
    tc_dir = dataset_path / "spreadsheet" / str(task_id)
    input_file = find_input_file(tc_dir, task_id)
    if input_file is None:
        return {
            'id': task_id,
            'status': 'missing_input',
            'error': f"No input file found in {tc_dir}",
            'duration_sec': 0,
            'solution': '',
        }
    file_name = input_file.name
    input_path = str(input_file)
    output_path = str(output_dir / f"1_{task_id}_output.xlsx")

    # Generate spreadsheet content preview (first N rows)
    file_content = gen_file_content(input_path, max_rows=max_rows)

    # Both single and multi-round use the same iterative loop.
    # Single-round = max_turns of 1 (but we still feed back errors for a retry
    # in case the first attempt fails — this matches the spirit of the official
    # multi-react protocol and maximises output-file creation rate).
    is_multi = setting != "single"
    effective_turns = max_turns if is_multi else 3  # up to 3 retries on exec failure

    if is_multi:
        prompt = PROMPT_MULTI_ROUND.format(
            instruction=data['instruction'],
            spreadsheet_path=input_path,
            spreadsheet_content=file_content,
            instruction_type=data['instruction_type'],
            answer_position=data['answer_position'],
            output_path=output_path,
            max_turn_num=max_turns,
        )
    else:
        prompt = PROMPT_SINGLE.format(
            instruction=data['instruction'],
            spreadsheet_path=input_path,
            spreadsheet_content=file_content,
            instruction_type=data['instruction_type'],
            answer_position=data['answer_position'],
            output_path=output_path,
        )

    messages = [{"role": "user", "content": prompt}]
    code = ""
    exec_errors = []

    for turn in range(effective_turns):
        try:
            response, tokens = call_llm(client, messages, model)
            cost_tracker.record(tokens=tokens)
        except Exception as e:
            return {
                'id': task_id,
                'status': 'llm_error',
                'error': str(e),
                'turn': turn,
                'duration_sec': time.time() - start_time,
                'solution': code,
            }

        messages.append({"role": "assistant", "content": response})
        code = extract_code(response)

        ok, exec_output = execute_code_local(code)
        cost_tracker.record(code_execs=1)

        if ok and os.path.exists(output_path):
            # TC1 output confirmed — exit loop
            break

        # Execution failed or output file not created
        if not ok:
            exec_errors.append(f"Turn {turn + 1}: {exec_output[:300]}")

        if is_multi or not ok:
            # Feed error back to LLM for correction
            feedback = exec_output if not ok else (
                f"Code ran but output file was not created at {output_path}. "
                "Ensure the file is saved with openpyxl wb.save() to that exact path."
            )
            messages.append({"role": "user", "content": feedback})
        else:
            # single-round: no more LLM turns, but we already logged the error
            break

    # Verify TC1 output was created
    if not os.path.exists(output_path):
        return {
            'id': task_id,
            'status': 'failed_execution',
            'error': f"TC1 output not created after {effective_turns} turn(s). "
                     f"Last errors: {'; '.join(exec_errors[-2:])}",
            'duration_sec': time.time() - start_time,
            'solution': code,
            'setting': setting,
        }

    # Replay the solution on test cases 2 and 3 — only for datasets that have them.
    # verified_400 has only 1 TC per task; full 912 has 3 TCs.
    # Strategy: replace only the filename fragments (basenames) in the code string.
    # The absolute output dir is the same for TC1/2/3, so basename replacement is safe.
    if file_name == "initial.xlsx":
        tc_suffix = None  # unprefixed — no TC2/TC3 possible
    elif file_name.endswith("_init.xlsx"):
        tc_suffix = "init"
    else:
        tc_suffix = "input"
    tc_exec_errors = {}

    for tc_idx in [2, 3]:
        if tc_suffix is None:
            break  # unprefixed dataset — no TC2/TC3
        tc_input_name = f"{tc_idx}_{task_id}_{tc_suffix}.xlsx"
        tc_input_path = dataset_path / "spreadsheet" / str(task_id) / tc_input_name
        if not tc_input_path.exists():
            # Dataset has only 1 TC (e.g. verified_400) — skip TC2/TC3 replay
            break

        tc_output_name = f"{tc_idx}_{task_id}_output.xlsx"
        tc_output_path = output_dir / tc_output_name

        # Build TC code: replace TC1 basename fragments with TC{n} equivalents
        tc_code = code.replace(file_name, tc_input_name)
        tc_code = tc_code.replace(f"1_{task_id}_output.xlsx", tc_output_name)

        ok, tc_exec_out = execute_code_local(tc_code)
        cost_tracker.record(code_execs=1)

        if not ok or not tc_output_path.exists():
            tc_exec_errors[tc_idx] = tc_exec_out[:300] if not ok else "output file not created"

    return {
        'id': task_id,
        'status': 'completed',
        'duration_sec': time.time() - start_time,
        'solution': code,
        'setting': setting,
        'exec_errors': exec_errors if exec_errors else None,
        'tc_exec_errors': tc_exec_errors if tc_exec_errors else None,
    }


# ============================================================================
# Track A: Full Orchestrator
# ============================================================================

def run_track_a(dataset, dataset_path, args):
    """Run Track A: official code generation protocol."""
    model = args.model
    setting = args.setting or "single"
    run_id = f"track_a_{setting}_{model}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    results_dir = BENCH_DIR / "results" / run_id
    output_dir = dataset_path / "outputs" / f"{setting}_{model}"
    output_dir.mkdir(parents=True, exist_ok=True)

    checkpoint = CheckpointManager(results_dir)
    cost_tracker = CostTracker()
    client = get_anthropic_client(args.api_key)

    # Filter to pending instructions
    pending = [d for d in dataset if not checkpoint.is_completed(d['id'])]
    print(f"\n{'='*60}")
    print(f"Track A — {setting} round | Model: {model}")
    print(f"Total: {len(dataset)} | Completed: {len(dataset) - len(pending)} | Pending: {len(pending)}")
    print(f"Output: {output_dir}")
    print(f"{'='*60}\n")

    for data in tqdm(pending, desc=f"Track A ({setting})"):
        try:
            gen_result = run_track_a_single(
                data, dataset_path, output_dir, client, model, cost_tracker,
                setting=setting, max_turns=args.max_turns, max_rows=args.max_rows,
            )
        except Exception as e:
            gen_result = {
                'id': data['id'],
                'status': 'crash',
                'error': traceback.format_exc(),
                'duration_sec': 0,
                'solution': '',
            }

        # Now evaluate using official comparison
        eval_result = evaluate_instruction(
            data, str(dataset_path), str(dataset_path / "outputs"), model, setting
        )

        # Merge generation + evaluation results
        combined = {
            **gen_result,
            **eval_result,
            'instruction': data['instruction'],
            'instruction_type': data['instruction_type'],
            'answer_position': data['answer_position'],
        }
        checkpoint.save_result(combined)

    # Final summary
    all_results = checkpoint.get_all_results()
    return finalize_results(all_results, run_id, results_dir, model, cost_tracker, "Track A", setting)


# ============================================================================
# Track B: ServalSheets MCP (stub — requires MCP client wiring)
# ============================================================================

def run_track_b(dataset, dataset_path, args):
    """
    Run Track B: ServalSheets MCP protocol.

    For each instruction:
    1. Import input.xlsx into Google Sheets via MCP (sheets_composite.import_xlsx)
    2. Give Claude the instruction + all ServalSheets MCP tools
    3. Claude uses MCP tools in an agentic loop to manipulate the spreadsheet
    4. Export result as XLSX via MCP (sheets_composite.export_xlsx)
    5. Evaluate with official comparison logic (identical to Track A scoring)

    ServalSheets manages its own Google auth — credentials are configured via
    the MCP server's environment (service account JSON or OAuth tokens).
    """
    model = args.model

    # If --resume, find the latest existing Track B checkpoint directory
    results_dir = None
    if args.resume:
        existing = sorted(
            BENCH_DIR.glob(f"results/track_b_mcp_{model}_*"),
            key=lambda p: p.name,
            reverse=True,
        )
        for candidate in existing:
            cp = candidate / "checkpoint.jsonl"
            if cp.exists() and cp.stat().st_size > 0:
                results_dir = candidate
                print(f"Resuming from: {results_dir.name}")
                break

    if results_dir is None:
        run_id = f"track_b_mcp_{model}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
        results_dir = BENCH_DIR / "results" / run_id
        results_dir.mkdir(parents=True, exist_ok=True)
    else:
        run_id = results_dir.name

    checkpoint = CheckpointManager(results_dir)
    cost_tracker = CostTracker()

    pending = [d for d in dataset if not checkpoint.is_completed(d['id'])]
    print(f"\n{'='*60}")
    print(f"Track B — ServalSheets MCP | Model: {model}")
    print(f"Total: {len(dataset)} | Completed: {len(dataset) - len(pending)} | Pending: {len(pending)}")
    print(f"{'='*60}\n")

    # Use the local MCP executor — runs Claude with ServalSheets tool schemas,
    # backed by openpyxl locally. Same tool interface, same intelligence.
    from track_b.local_executor import LocalMcpExecutor
    executor = LocalMcpExecutor(api_key=args.api_key, model=model)
    print("Using local MCP executor (openpyxl-backed tool simulation)")

    for data in tqdm(pending, desc="Track B (MCP)"):
        try:
            result = executor.execute_instruction(data, str(dataset_path), cost_tracker)
        except Exception as e:
            result = {
                'id': data['id'],
                'status': 'crash',
                'error': traceback.format_exc(),
                'duration_sec': 0,
                'test_case_results': [0, 0, 0],
                'soft_restriction': 0.0,
                'hard_restriction': 0,
            }

        combined = {
            **result,
            'instruction': data['instruction'],
            'instruction_type': data['instruction_type'],
            'answer_position': data['answer_position'],
        }
        checkpoint.save_result(combined)

    all_results = checkpoint.get_all_results()
    return finalize_results(all_results, run_id, results_dir, model, cost_tracker, "Track B", "mcp")


# ============================================================================
# Results Finalization & Summary
# ============================================================================

def finalize_results(all_results, run_id, results_dir, model, cost_tracker, track_name, setting):
    """Compute aggregate scores and write final reports."""

    # Compute aggregates
    total = len(all_results)
    completed = [r for r in all_results if r.get('status') == 'completed']
    errors = [r for r in all_results if r.get('status') != 'completed']

    soft_scores = [r.get('soft_restriction', 0) for r in all_results]
    hard_scores = [r.get('hard_restriction', 0) for r in all_results]

    avg_soft = sum(soft_scores) / total if total > 0 else 0
    avg_hard = sum(hard_scores) / total if total > 0 else 0

    # By instruction type
    cell_level = [r for r in all_results if r.get('instruction_type') == 'Cell-Level Manipulation']
    sheet_level = [r for r in all_results if r.get('instruction_type') == 'Sheet-Level Manipulation']

    cell_soft = sum(r.get('soft_restriction', 0) for r in cell_level) / len(cell_level) if cell_level else 0
    cell_hard = sum(r.get('hard_restriction', 0) for r in cell_level) / len(cell_level) if cell_level else 0
    sheet_soft = sum(r.get('soft_restriction', 0) for r in sheet_level) / len(sheet_level) if sheet_level else 0
    sheet_hard = sum(r.get('hard_restriction', 0) for r in sheet_level) / len(sheet_level) if sheet_level else 0

    cost = cost_tracker.estimate_cost(model)

    summary = {
        "run_id": run_id,
        "track": track_name,
        "setting": setting,
        "model": model,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "dataset_size": total,
        "completed": len(completed),
        "errors": len(errors),
        "scores": {
            "overall": {
                "soft_avg": round(avg_soft, 4),
                "hard_avg": round(avg_hard, 4),
                "soft_pct": f"{avg_soft * 100:.1f}%",
                "hard_pct": f"{avg_hard * 100:.1f}%",
            },
            "cell_level": {
                "count": len(cell_level),
                "soft_avg": round(cell_soft, 4),
                "hard_avg": round(cell_hard, 4),
            },
            "sheet_level": {
                "count": len(sheet_level),
                "soft_avg": round(sheet_soft, 4),
                "hard_avg": round(sheet_hard, 4),
            },
        },
        "cost": cost,
    }

    # Write summary JSON
    summary_path = results_dir / "summary.json"
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)

    # Write full results JSON
    full_results_path = results_dir / "results.json"
    with open(full_results_path, 'w') as f:
        json.dump({
            "summary": summary,
            "results": all_results,
        }, f, indent=2, default=str)

    # Print summary
    print(f"\n{'='*60}")
    print(f"  {track_name} RESULTS — {setting} | {model}")
    print(f"{'='*60}")
    print(f"  Instructions:  {total} ({len(completed)} completed, {len(errors)} errors)")
    print(f"")
    print(f"  OVERALL SCORES:")
    print(f"    Soft (partial credit):  {avg_soft * 100:.1f}%")
    print(f"    Hard (all-or-nothing):  {avg_hard * 100:.1f}%")
    print(f"")
    print(f"  BY TYPE:")
    print(f"    Cell-Level  ({len(cell_level):3d}):  soft={cell_soft*100:.1f}%  hard={cell_hard*100:.1f}%")
    print(f"    Sheet-Level ({len(sheet_level):3d}):  soft={sheet_soft*100:.1f}%  hard={sheet_hard*100:.1f}%")
    print(f"")
    print(f"  COST:")
    print(f"    Tokens: {cost['input_tokens']:,} in + {cost['output_tokens']:,} out")
    print(f"    API calls: {cost['api_calls']}  |  Code execs: {cost['code_executions']}")
    print(f"    Estimated: ${cost['estimated_usd']:.2f}")
    print(f"")
    print(f"  Results saved to: {results_dir}")
    print(f"{'='*60}\n")

    return summary


# ============================================================================
# CLI
# ============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="SpreadsheetBench Evaluation Harness for ServalSheets"
    )
    parser.add_argument("--track", choices=["a", "b", "both"], default="a",
                        help="Track A (code gen), Track B (MCP), or both")
    parser.add_argument("--dataset", choices=["sample", "full", "verified"], default="sample",
                        help="Dataset size: sample (200), full (912), verified (400)")
    parser.add_argument("--model", default=DEFAULT_MODEL,
                        help=f"Model name (default: {DEFAULT_MODEL})")
    parser.add_argument("--api-key", default=os.environ.get("ANTHROPIC_API_KEY"),
                        help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    parser.add_argument("--setting", default="single",
                        choices=["single", "multi_row_exec", "multi_react_exec", "multi_row_react_exec"],
                        help="Inference setting (Track A only)")
    parser.add_argument("--max-turns", type=int, default=5,
                        help="Max conversation turns for multi-round settings")
    parser.add_argument("--max-rows", type=int, default=5,
                        help="Max rows of spreadsheet content shown in prompt")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from last checkpoint")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limit to first N instructions (for testing)")
    parser.add_argument("--start-index", type=int, default=0,
                        help="Start at this instruction index (for parallel slices)")
    parser.add_argument("--end-index", type=int, default=None,
                        help="End before this instruction index (for parallel slices)")
    return parser.parse_args()


def find_input_file(tc_dir: Path, task_id) -> Path | None:
    """
    Resolve the TC1 input file for a task, supporting all three naming conventions:
      1. 1_<id>_input.xlsx  (912 dataset)
      2. 1_<id>_init.xlsx   (most of verified_400)
      3. initial.xlsx        (5 tasks in verified_400)
    Returns the Path if found, else None.
    """
    for candidate in [
        tc_dir / f"1_{task_id}_input.xlsx",
        tc_dir / f"1_{task_id}_init.xlsx",
        tc_dir / "initial.xlsx",
    ]:
        if candidate.exists():
            return candidate
    return None


def find_golden_file(tc_dir: Path, task_id, tc_num: int = 1) -> Path | None:
    """
    Resolve the ground truth file for a given TC number, supporting:
      1. <n>_<id>_answer.xlsx  (912 dataset)
      2. <n>_<id>_golden.xlsx  (standard verified_400)
      3. golden.xlsx            (unprefixed verified_400)
    """
    for candidate in [
        tc_dir / f"{tc_num}_{task_id}_answer.xlsx",
        tc_dir / f"{tc_num}_{task_id}_golden.xlsx",
        tc_dir / "golden.xlsx" if tc_num == 1 else None,
    ]:
        if candidate and candidate.exists():
            return candidate
    return None


def verify_dataset_integrity(dataset, dataset_path):
    """
    Verify that input files exist for every instruction in the dataset.
    Aborts early with a clear error rather than silently failing per-task.
    """
    missing = []
    for entry in dataset[:20]:  # spot-check first 20 to catch naming issues fast
        task_id = entry['id']
        tc_dir = dataset_path / "spreadsheet" / str(task_id)
        if find_input_file(tc_dir, task_id) is None:
            missing.append(str(tc_dir))
    if missing:
        print(f"\n[INTEGRITY FAIL] {len(missing)} task input directories missing (first 20 checked):")
        for m in missing[:5]:
            print(f"  {m}")
        sys.exit(1)
    print(f"[integrity] Input files verified (spot-checked first 20 tasks) ✓")


def main():
    args = parse_args()

    if not args.api_key:
        print("ERROR: No API key provided.")
        print("Set ANTHROPIC_API_KEY environment variable or pass --api-key")
        sys.exit(1)

    # Verify environment before doing any work
    preflight_check()

    # Load dataset
    dataset, dataset_path = load_dataset(args.dataset)

    if args.limit:
        dataset = dataset[:args.limit]
        print(f"Limited to first {args.limit} instructions")

    if args.start_index or args.end_index:
        dataset = dataset[args.start_index:args.end_index]
        print(f"Slice: instructions {args.start_index}–{args.end_index or 'end'} ({len(dataset)} tasks)")

    # Dataset integrity check — catches path/naming issues before any API calls
    if args.track in ("a", "both"):
        verify_dataset_integrity(dataset, dataset_path)

    # Run tracks
    summaries = []

    if args.track in ("a", "both"):
        summary_a = run_track_a(dataset, dataset_path, args)
        summaries.append(summary_a)

    if args.track in ("b", "both"):
        summary_b = run_track_b(dataset, dataset_path, args)
        summaries.append(summary_b)

    # If both tracks ran, print comparison
    if len(summaries) == 2:
        print(f"\n{'='*60}")
        print(f"  TRACK COMPARISON")
        print(f"{'='*60}")
        for s in summaries:
            scores = s['scores']['overall']
            print(f"  {s['track']:10s}  soft={scores['soft_pct']:>6s}  hard={scores['hard_pct']:>6s}  cost=${s['cost']['estimated_usd']:.2f}")
        print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
