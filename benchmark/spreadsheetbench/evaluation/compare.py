"""
Cell-level comparison engine — EXACT PORT of SpreadsheetBench's official evaluation.py.

This module reproduces the official scoring logic byte-for-byte so our results
are directly comparable to published benchmarks.

Source: https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/evaluation/evaluation.py
"""

import datetime
import os
import openpyxl


def datetime_to_float(dt):
    """Convert datetime to Excel serial number (days since 1899-12-30)."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """
    Normalize a cell value for comparison.

    Official rules:
    - int/float → round(float(v), 2)
    - datetime.time → str(v)[:-3]  (strip seconds)
    - datetime.datetime → round(excel_serial, 0)
    - str → try parse as float and round(2), else keep string
    """
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    """
    Compare two cell values using official SpreadsheetBench rules.

    - Empty string == None (both directions)
    - Type mismatch after transform → False
    - Exact match after transform → True
    """
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    if v1 == v2:
        return True
    else:
        return False


def col_num2name(n):
    """Convert a column number to an Excel column name (1-indexed)."""
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a column number (1-indexed)."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    """Parse 'A1:AB12' into ((col_num, row_num), (col_num, row_num))."""
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    """Generate all cell names in a range like 'A1:C3'."""
    if ':' not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    """Compare cells in a specific range between ground truth and processed workbooks."""
    if sheet_name not in wb_proc:
        return False, "worksheet not found", []
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)
    mismatches = []

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            mismatches.append({
                'cell': cell_name,
                'expected': repr(cell_gt.value),
                'actual': repr(cell_proc.value),
                'expected_transformed': repr(transform_value(cell_gt.value)),
                'actual_transformed': repr(transform_value(cell_proc.value)),
            })

    if mismatches:
        msg = f"Value difference at cell {mismatches[0]['cell']}: " \
              f"expected {mismatches[0]['expected']}, got {mismatches[0]['actual']}"
        return False, msg, mismatches

    return True, "", []


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    """
    Compare a ground-truth workbook against a processed (output) workbook.

    This is the TOP-LEVEL comparison function matching the official evaluation.py exactly.

    Returns: (passed: bool, message: str, details: list)
    """
    import os
    if not os.path.exists(proc_file):
        return False, "File not exist", []

    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e), []

    sheet_cell_ranges = answer_position.split(',')
    all_mismatches = []

    for sheet_cell_range in sheet_cell_ranges:
        if '!' in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split('!')
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        sheet_name = sheet_name.lstrip("'").rstrip("'")
        cell_range = cell_range.lstrip("'").rstrip("'")

        result, msg, mismatches = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        all_mismatches.extend(mismatches)
        if not result:
            return False, msg, all_mismatches

    return True, "", []


def _effective_answer_position(dataset_entry):
    """
    Build the answer_position string used for cell comparison.

    If the dataset entry includes an 'answer_sheet' field and the
    answer_position does not already reference a sheet (no '!' present),
    prepend the sheet name so the evaluator checks the correct sheet.
    """
    pos = dataset_entry['answer_position']
    sheet = dataset_entry.get('answer_sheet', '')
    if sheet and '!' not in pos:
        # Wrap sheet name in quotes if it contains spaces
        if ' ' in sheet:
            pos = f"'{sheet}'!{pos}"
        else:
            pos = f"{sheet}!{pos}"
    return pos


def _count_test_cases(dataset_path, task_id):
    """
    Detect how many ground-truth test cases exist for a task.
    Returns 3 for 912-style datasets, 1 for verified_400-style (including
    both prefixed and unprefixed naming conventions).
    """
    tc_dir = os.path.join(dataset_path, "spreadsheet", str(task_id))
    # Check for TC2 ground truth (answer or golden)
    for suffix in ("answer", "golden"):
        if os.path.exists(os.path.join(tc_dir, f"2_{task_id}_{suffix}.xlsx")):
            return 3
    return 1


def evaluate_instruction(dataset_entry, dataset_path, output_dir, model_name, setting="single"):
    """
    Evaluate a single instruction across available test cases.

    Automatically detects the number of test cases (1 for verified_400,
    3 for the full 912 dataset) to avoid false failures on TC2/TC3.

    Returns dict with:
    - id, instruction_type
    - test_case_results: list of 0/1 per TC
    - soft_restriction: float (fraction passed)
    - hard_restriction: 0 or 1
    - details: per-test-case mismatch info
    """
    import os

    task_id = dataset_entry['id']

    # Detect dataset variant
    num_tcs = _count_test_cases(dataset_path, task_id)

    # Build effective answer_position (inject sheet name if needed)
    answer_position = _effective_answer_position(dataset_entry)

    test_case_results = []
    test_case_details = []

    for tc_idx in range(num_tcs):
        # Support all naming conventions:
        #   <n>_<id>_answer.xlsx  (912 dataset)
        #   <n>_<id>_golden.xlsx  (standard verified_400)
        #   golden.xlsx            (unprefixed verified_400)
        _tc = tc_idx + 1
        _gt_dir = os.path.join(dataset_path, "spreadsheet", str(task_id))
        _answer_name = f"{_tc}_{task_id}_answer.xlsx"
        _golden_name = f"{_tc}_{task_id}_golden.xlsx"
        _plain_golden = "golden.xlsx"
        if os.path.exists(os.path.join(_gt_dir, _answer_name)):
            gt_path = os.path.join(_gt_dir, _answer_name)
        elif os.path.exists(os.path.join(_gt_dir, _golden_name)):
            gt_path = os.path.join(_gt_dir, _golden_name)
        else:
            gt_path = os.path.join(_gt_dir, _plain_golden)  # fallback
        proc_path = os.path.join(
            output_dir, f"{setting}_{model_name}",
            f"{_tc}_{task_id}_output.xlsx"
        )

        try:
            result, msg, mismatches = compare_workbooks(
                gt_path, proc_path,
                dataset_entry['instruction_type'],
                answer_position
            )
        except Exception as e:
            result = False
            msg = str(e)
            mismatches = []

        test_case_results.append(int(result))
        test_case_details.append({
            'test_case': _tc,
            'passed': result,
            'message': msg,
            'mismatch_count': len(mismatches),
            'first_mismatch': mismatches[0] if mismatches else None,
        })

    soft = test_case_results.count(1) / len(test_case_results)
    hard = 0 if 0 in test_case_results else 1

    return {
        'id': task_id,
        'instruction_type': dataset_entry['instruction_type'],
        'test_case_results': test_case_results,
        'soft_restriction': soft,
        'hard_restriction': hard,
        'num_test_cases': num_tcs,
        'details': test_case_details,
    }
